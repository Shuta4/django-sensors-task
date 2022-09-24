from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Prefetch
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework_extensions import mixins
from openpyxl import load_workbook, Workbook
from tempfile import NamedTemporaryFile
from uuid import UUID
from datetime import datetime
import logging
from .models import Sensor, Report
from .serializers import SensorSerializer, ReportSerializer


class SensorViewSet(mixins.NestedViewSetMixin, viewsets.ModelViewSet):
    queryset = Sensor.objects.all()
    serializer_class = SensorSerializer

    @action(detail=False, methods=['get'])
    def xlsx(self, request: Request) -> HttpResponse:
        """
        Get information about all sensors and their last reports in xlsx-file.
        """
        def excel_format(value):
            if isinstance(value, UUID):
                return str(value)
            if isinstance(value, datetime):
                return value.replace(tzinfo=None)
            return value

        ordered_reports = Report.objects.order_by('-datetime')
        sensors_query = Sensor.objects.prefetch_related(
            Prefetch('report_set', queryset=ordered_reports, to_attr='reports')
        )
        wb = Workbook()
        ws = wb.active

        sensor_fields = [f.name for f in Sensor._meta.get_fields() if f.name != 'report']
        report_fields = [f.name for f in Report._meta.get_fields() if f.name != 'sensor']

        for col, field_name in enumerate([f"sensor_{x}" for x in sensor_fields] + report_fields, start=1):
            ws.cell(row=1, column=col).value = field_name

        for row, sensor in enumerate(sensors_query, start=2):
            count = 0
            for col, field_name in enumerate(sensor_fields, start=1):
                ws.cell(row=row, column=col).value = excel_format(getattr(sensor, field_name))
                count += 1

            if len(sensor.reports) != 0:
                report = sensor.reports[0]
                for col, field_name in enumerate(report_fields, start=count + 1):
                    ws.cell(row=row, column=col).value = excel_format(getattr(report, field_name))

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            data = tmp.read()

        res = HttpResponse(content_type="application/ms-excel")
        res['Content-Disposition'] = 'attachment; filename="export.xlsx"'
        res.write(data)
        return res

    @xlsx.mapping.post
    def import_xlsx(self, request: Request) -> Response:
        """
        Import list of sensors from xlsx file.
        """
        for file_key in request.FILES:
            wb = load_workbook(request.FILES[file_key].file,
                               read_only=True, data_only=True)
            ws = wb.active
            columns = {
                'id': None,
                'address': None,
                'lat': None,
                'lon': None,
                'iccid': None,
                'stock_number': None,
            }
            with transaction.atomic():
                for row in ws.values:
                    if columns['id'] is None:
                        for i, value in enumerate(row):
                            if value == 'sensor_uuid':
                                columns['id'] = i
                            else:
                                columns[value] = i
                        if columns['id'] is None:
                            return Response(
                                {'detail': "File does not contain mandatory column 'sensor_uuid'"},
                                status=status.HTTP_400_BAD_REQUEST
                            )
                        continue
                    sensor = Sensor()
                    for key, i in columns.items():
                        if key is None or i is None:
                            continue
                        if row[i] is None and (key == 'iccid' or key == 'address' or key == 'stock_number'):
                            value = ''
                        elif row[i] == '' and (key == 'lat' or key == 'lon'):
                            value = None
                        else:
                            value = row[i]
                        setattr(sensor, key, value)

                    try:
                        sensor.save()
                    except ValidationError as e:
                        logging.warning(sensor.id, e)

            break
        return Response(status=status.HTTP_200_OK)


class ReportViewSet(mixins.NestedViewSetMixin, viewsets.ModelViewSet):
    queryset = Report.objects.all()
    serializer_class = ReportSerializer
